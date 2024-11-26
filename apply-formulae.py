from openpyxl.worksheet import worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load the workbook and select the active worksheet
wb = load_workbook('requirement/sales_record.xlsx')
ws: worksheet = wb['record']

# Set the minimum and maximum columns and rows to work with
min_column = ws.min_column
max_column = ws.max_column
min_row = ws.min_row
max_row = ws.max_row

# Iterate through each column from min_column to max_column
for i in range(min_column, max_column + 1):
    # Get the column letter using the get_column_letter function
    letter = get_column_letter(i)
    # Apply the SUM formula to the cell below the last row in the current column
    ws[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    # Apply the 'Currency' style to the cell with the SUM formula
    ws[f'{letter}{max_row + 1}'].style = 'Currency'

# Save the workbook with the changes
wb.save('requirement/my_report.xlsx')
